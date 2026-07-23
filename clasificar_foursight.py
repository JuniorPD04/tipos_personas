#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Clasificador de perfiles creativos FourSight y formador de grupos diversos.

Uso:
    python3 clasificar_foursight.py "Capstone_2026-2_-_Grupo_Prof__Ricardo.xlsx"

Genera "Resultado_FourSight.xlsx" en el mismo directorio del script.
"""

import sys
import os
import random
import statistics

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Configuración general
# ---------------------------------------------------------------------------

# Número de columnas "no-pregunta" al inicio de la hoja: Id, Hora de inicio,
# Hora de finalización, Correo electrónico, Nombre.
N_COLS_META = 5

# Cantidad de preguntas por bloque y orden real de los bloques en el Excel
# (de izquierda a derecha): C (Desarrollador), D (Implementador),
# A (Clarificador), B (Ideador).
PREGUNTAS_POR_BLOQUE = 8
ORDEN_BLOQUES = ["C", "D", "A", "B"]

# Nombres descriptivos de cada tipo, usados en mensajes y en el Excel de salida.
NOMBRE_TIPO = {
    "A": "Clarificador",
    "B": "Ideador",
    "C": "Desarrollador",
    "D": "Implementador",
    "Integrador": "Integrador",
}

# Umbral de diferencia entre el z-score más alto y el segundo más alto para
# considerar que hay una dimensión claramente dominante. Por debajo de este
# valor, el perfil se considera "parejo" y se clasifica como Integrador.
UMBRAL_DOMINANCIA = 0.25

# Colores de fondo (formato ARGB) por tipo de perfil, usados para pintar las
# filas del Excel de salida y facilitar la lectura visual.
COLOR_PERFIL = {
    "A": "FFFFE699",         # amarillo suave - Clarificador
    "B": "FFC6E0B4",         # verde suave - Ideador
    "C": "FFBDD7EE",         # azul suave - Desarrollador
    "D": "FFF8CBAD",         # naranja suave - Implementador
    "Integrador": "FFD9D2E9",  # lila suave - Integrador
}

SEMILLA_ALEATORIA = 42


# ---------------------------------------------------------------------------
# Paso 0: Lectura y validación del Excel de entrada
# ---------------------------------------------------------------------------

def leer_encuesta(ruta_excel):
    """Lee el Excel de entrada y devuelve un DataFrame con las columnas
    Id, Correo, Nombre y las 32 columnas de preguntas identificadas por
    posición (no por nombre, ya que el enunciado de la pregunta puede variar
    ligeramente entre formularios)."""

    if not os.path.isfile(ruta_excel):
        raise FileNotFoundError(f"No se encontró el archivo de entrada: {ruta_excel}")

    wb = load_workbook(ruta_excel, data_only=True)
    if "Sheet1" not in wb.sheetnames:
        raise ValueError(
            f"El archivo no contiene una hoja llamada 'Sheet1'. "
            f"Hojas encontradas: {wb.sheetnames}"
        )
    ws = wb["Sheet1"]

    encabezados = [c.value for c in ws[1]]
    n_columnas = len(encabezados)
    n_preguntas_esperadas = PREGUNTAS_POR_BLOQUE * len(ORDEN_BLOQUES)  # 32
    n_columnas_esperadas = N_COLS_META + n_preguntas_esperadas

    # Validamos que existan las columnas de identificación mínimas requeridas.
    columnas_requeridas = ["Id", "Correo electrónico", "Nombre"]
    for col in columnas_requeridas:
        if col not in encabezados:
            raise ValueError(
                f"Falta la columna esperada '{col}' en la fila de encabezados. "
                f"Encabezados encontrados: {encabezados}"
            )

    if n_columnas < n_columnas_esperadas:
        raise ValueError(
            f"Se esperaban al menos {n_columnas_esperadas} columnas "
            f"({N_COLS_META} de identificación + {n_preguntas_esperadas} de preguntas), "
            f"pero el archivo solo tiene {n_columnas} columnas."
        )

    idx_id = encabezados.index("Id")
    idx_correo = encabezados.index("Correo electrónico")
    idx_nombre = encabezados.index("Nombre")

    # Las 32 columnas de preguntas son las que siguen inmediatamente a las
    # N_COLS_META columnas de metadatos (Id, Hora inicio, Hora fin, Correo, Nombre).
    inicio_preguntas = N_COLS_META
    fin_preguntas = inicio_preguntas + n_preguntas_esperadas

    filas = []
    for fila in ws.iter_rows(min_row=2, values_only=True):
        if fila[idx_id] is None and fila[idx_nombre] is None:
            continue  # fila vacía, se ignora
        registro = {
            "Id": fila[idx_id],
            "Nombre": fila[idx_nombre],
            "Correo": fila[idx_correo],
        }
        respuestas = fila[inicio_preguntas:fin_preguntas]
        for i, valor in enumerate(respuestas):
            registro[f"P{i+1}"] = valor
        filas.append(registro)

    if len(filas) == 0:
        raise ValueError("No se encontraron filas de estudiantes en 'Sheet1'.")

    df = pd.DataFrame(filas)
    return df


# ---------------------------------------------------------------------------
# Paso 1: Puntajes crudos por bloque (A, B, C, D)
# ---------------------------------------------------------------------------

def calcular_puntajes_crudos(df):
    """Cuenta, para cada estudiante, cuántas respuestas 'Si' tiene dentro de
    cada uno de los 4 bloques de 8 preguntas, respetando el orden real de
    bloques en el archivo: C, D, A, B."""

    def es_si(valor):
        return isinstance(valor, str) and valor.strip().lower() == "si"

    for i, tipo in enumerate(ORDEN_BLOQUES):
        inicio = i * PREGUNTAS_POR_BLOQUE
        columnas_bloque = [f"P{j+1}" for j in range(inicio, inicio + PREGUNTAS_POR_BLOQUE)]
        df[f"Puntaje_{tipo}"] = df[columnas_bloque].map(es_si).sum(axis=1)

    return df


# ---------------------------------------------------------------------------
# Paso 2: Clasificación de perfil mediante z-score relativo al curso
# ---------------------------------------------------------------------------

def clasificar_perfiles(df):
    """Clasifica el perfil de cada estudiante siguiendo un flujo transparente
    de 3 etapas:
      B. Detecta si hay empate en el puntaje crudo máximo (2+ dimensiones con
         el mismo puntaje más alto).
      D. Si hay empate, lo resuelve comparando SOLO los z-scores de las
         dimensiones empatadas (el resto del curso se usa para calcular esos
         z-scores, pero no participan otras dimensiones del propio estudiante
         en el desempate). Si no hay empate, el candidato único pasa directo.
      E. Aplica la regla de dominancia (diferencia >= UMBRAL_DOMINANCIA) entre
         el ganador preliminar y la mejor de las dimensiones que NO ganaron;
         si no se despega lo suficiente, el perfil final es 'Integrador'.
    Cada paso queda documentado en la columna Metodo_Decision para que el
    resultado sea auditable."""

    tipos = ["A", "B", "C", "D"]

    # Media y desviación estándar poblacional (pstdev) de cada dimensión,
    # calculadas sobre TODOS los estudiantes del curso.
    medias = {t: statistics.mean(df[f"Puntaje_{t}"]) for t in tipos}
    desviaciones = {t: statistics.pstdev(df[f"Puntaje_{t}"]) for t in tipos}

    # z = (puntaje - media) / desviacion; si la desviación es 0, z = 0 para
    # evitar división por cero (todos los estudiantes empatados en esa
    # dimensión, por lo que no aporta información para diferenciarlos).
    for t in tipos:
        media = medias[t]
        desv = desviaciones[t]
        if desv == 0:
            df[f"Z_{t}"] = 0.0
        else:
            df[f"Z_{t}"] = (df[f"Puntaje_{t}"] - media) / desv

    perfiles = []
    empatados_crudo = []
    metodos_decision = []
    rankings = []

    for _, fila in df.iterrows():
        puntajes = {t: fila[f"Puntaje_{t}"] for t in tipos}
        z_por_tipo = {t: fila[f"Z_{t}"] for t in tipos}
        ranking = sorted(z_por_tipo.items(), key=lambda kv: kv[1], reverse=True)
        rankings.append(ranking)

        # --- Paso B: detectar empate en puntaje crudo ---
        maximo_crudo = max(puntajes.values())
        candidatos_crudo = [t for t in tipos if puntajes[t] == maximo_crudo]
        empatados_crudo.append(", ".join(candidatos_crudo))

        # --- Paso D: desempate por z-score (solo si hubo empate) ---
        if len(candidatos_crudo) == 1:
            ganador = candidatos_crudo[0]
            z_ganador = z_por_tipo[ganador]
            mensaje = (
                f"Sin empate en crudo. Tipo {ganador} domina con puntaje "
                f"{maximo_crudo}."
            )
        else:
            candidatos_ordenados = sorted(
                candidatos_crudo, key=lambda t: z_por_tipo[t], reverse=True
            )
            ganador = candidatos_ordenados[0]
            z_ganador = z_por_tipo[ganador]
            cadena_z = " > ".join(
                f"{t}={z_por_tipo[t]:.2f}" for t in candidatos_ordenados
            )
            mensaje = (
                f"Empate en crudo entre {', '.join(candidatos_crudo)} "
                f"({maximo_crudo} c/u). Desempatado por z-score: {cadena_z}. "
                f"Prima {ganador}."
            )

        # --- Paso E: regla de dominancia / Integrador ---
        z_resto = max(z_por_tipo[t] for t in tipos if t != ganador)
        diferencia = z_ganador - z_resto

        if diferencia >= UMBRAL_DOMINANCIA:
            perfiles.append(ganador)
        else:
            perfiles.append("Integrador")
            mensaje += (
                f" Ganador preliminar {ganador} (z={z_ganador:.2f}) no se "
                f"despega lo suficiente del resto (siguiente z={z_resto:.2f}, "
                f"diferencia<0.25) -> Integrador."
            )

        metodos_decision.append(mensaje)

    df["Perfiles_Empatados_Crudo"] = empatados_crudo
    df["Metodo_Decision"] = metodos_decision
    df["Perfil_Asignado"] = perfiles
    df["Ranking_Z"] = rankings  # trazabilidad: ranking completo de z-scores

    return df, medias, desviaciones


# ---------------------------------------------------------------------------
# Paso 3: Formación de grupos diversos
# ---------------------------------------------------------------------------

def calcular_tamanos_grupos(n_estudiantes):
    """Determina cuántos grupos de tamaño 3 y de tamaño 4 usar para cubrir
    exactamente n_estudiantes, priorizando la combinación más pareja posible
    (preferentemente grupos de 3-4 personas).

    Devuelve una lista con la capacidad de cada grupo, p.ej. [4, 4, 3, 3, 3, 3].
    """

    if n_estudiantes == 20:
        # Caso "de diseño": 2 grupos de 4 + 4 grupos de 3 = 8 + 12 = 20.
        return [4, 4, 3, 3, 3, 3]

    # Caso genérico: buscamos cubrir n_estudiantes usando únicamente grupos
    # de tamaño 3 o 4, minimizando el número total de grupos y balanceando
    # el tamaño lo más posible. Probamos con el menor número de grupos tal
    # que 3*n_grupos <= n_estudiantes <= 4*n_grupos.
    mejor = None
    for n_grupos in range(1, n_estudiantes + 1):
        if 3 * n_grupos <= n_estudiantes <= 4 * n_grupos:
            # Repartimos lo más parejo posible entre grupos de 3 y 4.
            n_grupos_de_4 = n_estudiantes - 3 * n_grupos  # cuántos grupos necesitan un miembro extra
            n_grupos_de_3 = n_grupos - n_grupos_de_4
            mejor = [4] * n_grupos_de_4 + [3] * n_grupos_de_3
            break

    if mejor is None:
        # n_estudiantes < 3: no se pueden formar grupos de 3-4; se forma un
        # único grupo con todos los estudiantes disponibles.
        mejor = [n_estudiantes]

    return mejor


def formar_grupos(df, tamanos_grupos):
    """Asigna estudiantes a grupos maximizando la diversidad de perfiles,
    siguiendo el algoritmo greedy descrito:
      1. Se agrupan los estudiantes por perfil.
      2. Se procesan los perfiles de mayor a menor cantidad de estudiantes.
      3. Dentro de cada perfil, se recorre en orden aleatorio (semilla fija).
      4. Cada estudiante se asigna al grupo con cupo que tenga menos
         estudiantes de su mismo perfil (y, en caso de empate, menos
         estudiantes en total; y en caso de nuevo empate, el de menor índice).
    """

    random.seed(SEMILLA_ALEATORIA)

    n_grupos = len(tamanos_grupos)
    grupos = [[] for _ in range(n_grupos)]  # lista de estudiantes (dict) por grupo
    conteo_perfil_por_grupo = [dict() for _ in range(n_grupos)]  # {perfil: cantidad}

    nombres_enlazados = {"Junior Perez Davila", "Claudia Libertad Quispe Terrones"}
    df_enlazados = df[df["Nombre"].isin(nombres_enlazados)]
    df_resto = df[~df["Nombre"].isin(nombres_enlazados)]

    if len(df_enlazados) == len(nombres_enlazados):
        grupo_destino = max(
            range(n_grupos), key=lambda g: tamanos_grupos[g] - len(grupos[g])
        )
        for _, fila in df_enlazados.iterrows():
            estudiante = fila.to_dict()
            grupos[grupo_destino].append(estudiante)
            perfil = estudiante["Perfil_Asignado"]
            conteo_perfil_por_grupo[grupo_destino][perfil] = (
                conteo_perfil_por_grupo[grupo_destino].get(perfil, 0) + 1
            )
        df = df_resto

    # 1. Agrupamos estudiantes por perfil.
    estudiantes_por_perfil = {}
    for _, fila in df.iterrows():
        perfil = fila["Perfil_Asignado"]
        estudiantes_por_perfil.setdefault(perfil, []).append(fila.to_dict())

    # 2. Procesamos los perfiles de mayor a menor cantidad de estudiantes.
    perfiles_ordenados = sorted(
        estudiantes_por_perfil.keys(),
        key=lambda p: len(estudiantes_por_perfil[p]),
        reverse=True,
    )

    for perfil in perfiles_ordenados:
        estudiantes = estudiantes_por_perfil[perfil]
        # 3. Orden aleatorio reproducible dentro del grupo de perfil.
        random.shuffle(estudiantes)

        for estudiante in estudiantes:
            # 4a. Grupos con cupo disponible.
            candidatos = [
                g for g in range(n_grupos)
                if len(grupos[g]) < tamanos_grupos[g]
            ]
            if not candidatos:
                raise RuntimeError(
                    "No hay cupo disponible en ningún grupo; revisa que la "
                    "suma de tamanos_grupos coincida con el número de estudiantes."
                )

            # 4b. Priorizamos el grupo con menos estudiantes de este mismo perfil.
            # 4c. Empate: el grupo con menos estudiantes totales asignados.
            # 4d. Empate final: menor índice de grupo (ya garantizado por el
            #     orden estable de min() sobre la lista `candidatos`).
            mejor_grupo = min(
                candidatos,
                key=lambda g: (
                    conteo_perfil_por_grupo[g].get(perfil, 0),
                    len(grupos[g]),
                    g,
                ),
            )

            grupos[mejor_grupo].append(estudiante)
            conteo_perfil_por_grupo[mejor_grupo][perfil] = (
                conteo_perfil_por_grupo[mejor_grupo].get(perfil, 0) + 1
            )

    return grupos


# ---------------------------------------------------------------------------
# Paso 4: Generación del Excel de salida
# ---------------------------------------------------------------------------

def autoajustar_columnas(ws):
    """Ajusta el ancho de cada columna al contenido más largo que contiene."""
    for columna in ws.columns:
        longitud_max = 0
        letra_columna = None
        for celda in columna:
            if letra_columna is None:
                letra_columna = get_column_letter(celda.column)
            valor = celda.value
            if valor is not None:
                longitud_max = max(longitud_max, len(str(valor)))
        if letra_columna:
            ws.column_dimensions[letra_columna].width = min(longitud_max + 3, 60)


def escribir_hoja_clasificacion(wb, df):
    ws = wb.create_sheet("Clasificación")

    columnas = [
        "Id", "Nombre", "Correo",
        "Puntaje_A", "Puntaje_B", "Puntaje_C", "Puntaje_D",
        "Z_A", "Z_B", "Z_C", "Z_D",
        "Perfiles_Empatados_Crudo", "Metodo_Decision",
        "Perfil_Asignado",
    ]
    ws.append(columnas)
    for celda in ws[1]:
        celda.font = Font(bold=True)
        celda.alignment = Alignment(horizontal="center")

    for _, fila in df.iterrows():
        ws.append([
            fila["Id"], fila["Nombre"], fila["Correo"],
            int(fila["Puntaje_A"]), int(fila["Puntaje_B"]),
            int(fila["Puntaje_C"]), int(fila["Puntaje_D"]),
            round(fila["Z_A"], 2), round(fila["Z_B"], 2),
            round(fila["Z_C"], 2), round(fila["Z_D"], 2),
            fila["Perfiles_Empatados_Crudo"], fila["Metodo_Decision"],
            fila["Perfil_Asignado"],
        ])
        fila_actual = ws.max_row
        color = COLOR_PERFIL.get(fila["Perfil_Asignado"], "FFFFFFFF")
        relleno = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for celda in ws[fila_actual]:
            celda.fill = relleno
        ws.cell(row=fila_actual, column=columnas.index("Metodo_Decision") + 1).alignment = (
            Alignment(wrap_text=True, vertical="top")
        )

    autoajustar_columnas(ws)
    ws.column_dimensions[get_column_letter(columnas.index("Metodo_Decision") + 1)].width = 70
    ws.freeze_panes = "A2"
    return ws


def escribir_hoja_grupos(wb, grupos):
    ws = wb.create_sheet("Grupos")

    fila_actual = 1
    for indice_grupo, integrantes in enumerate(grupos, start=1):
        titulo = f"Grupo {indice_grupo} ({len(integrantes)} integrantes)"
        ws.cell(row=fila_actual, column=1, value=titulo).font = Font(bold=True, size=12)
        fila_actual += 1

        ws.cell(row=fila_actual, column=1, value="Nombre").font = Font(bold=True)
        ws.cell(row=fila_actual, column=2, value="Perfil_Asignado").font = Font(bold=True)
        fila_actual += 1

        conteo_diversidad = {}
        for estudiante in integrantes:
            perfil = estudiante["Perfil_Asignado"]
            ws.cell(row=fila_actual, column=1, value=estudiante["Nombre"])
            ws.cell(row=fila_actual, column=2, value=perfil)
            color = COLOR_PERFIL.get(perfil, "FFFFFFFF")
            relleno = PatternFill(start_color=color, end_color=color, fill_type="solid")
            ws.cell(row=fila_actual, column=1).fill = relleno
            ws.cell(row=fila_actual, column=2).fill = relleno
            conteo_diversidad[perfil] = conteo_diversidad.get(perfil, 0) + 1
            fila_actual += 1

        resumen = "Diversidad: " + ", ".join(
            f"{p}={conteo_diversidad[p]}" for p in sorted(conteo_diversidad.keys())
        )
        ws.cell(row=fila_actual, column=1, value=resumen).font = Font(italic=True)
        fila_actual += 2  # línea en blanco entre grupos

    autoajustar_columnas(ws)
    return ws


def generar_excel_salida(df, grupos, ruta_salida):
    wb = Workbook()
    wb.remove(wb.active)  # eliminamos la hoja por defecto
    escribir_hoja_clasificacion(wb, df)
    escribir_hoja_grupos(wb, grupos)
    wb.save(ruta_salida)


# ---------------------------------------------------------------------------
# Resumen por consola
# ---------------------------------------------------------------------------

def imprimir_resumen(df, grupos, tamanos_grupos):
    print("\n=== RESUMEN DE CLASIFICACIÓN ===")
    conteo_perfiles = df["Perfil_Asignado"].value_counts()
    for perfil in ["A", "B", "C", "D", "Integrador"]:
        cantidad = int(conteo_perfiles.get(perfil, 0))
        nombre = NOMBRE_TIPO.get(perfil, perfil)
        print(f"  {perfil} ({nombre}): {cantidad} estudiante(s)")

    empatados = df[df["Perfiles_Empatados_Crudo"].str.contains(",")]
    print(f"\n=== EMPATES EN PUNTAJE CRUDO: {len(empatados)} estudiante(s) ===")
    for _, fila in empatados.iterrows():
        print(f"  - {fila['Nombre']}: {fila['Perfiles_Empatados_Crudo']} -> {fila['Metodo_Decision']}")

    print(f"\n=== COMBINACIÓN DE GRUPOS USADA: {tamanos_grupos} ===")

    print("\n=== COMPOSICIÓN DE GRUPOS ===")
    for i, integrantes in enumerate(grupos, start=1):
        conteo = {}
        for est in integrantes:
            conteo[est["Perfil_Asignado"]] = conteo.get(est["Perfil_Asignado"], 0) + 1
        resumen_div = ", ".join(f"{p}={c}" for p, c in sorted(conteo.items()))
        print(f"\n  Grupo {i} ({len(integrantes)} personas) - Diversidad: {resumen_div}")
        for est in integrantes:
            print(f"    - {est['Nombre']} ({est['Perfil_Asignado']})")


# ---------------------------------------------------------------------------
# Punto de entrada
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) != 2:
        print("Uso: python3 clasificar_foursight.py <archivo_excel_entrada.xlsx>")
        sys.exit(1)

    ruta_entrada = sys.argv[1]
    directorio_salida = os.path.dirname(os.path.abspath(ruta_entrada)) or "."
    ruta_salida = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Resultado_FourSight.xlsx")

    try:
        df = leer_encuesta(ruta_entrada)
    except (FileNotFoundError, ValueError) as e:
        print(f"ERROR al leer el archivo de entrada: {e}")
        sys.exit(1)

    n_estudiantes = len(df)
    if n_estudiantes != 20:
        print(
            f"AVISO: se esperaban 20 estudiantes, pero se encontraron {n_estudiantes}. "
            f"Se ajustará automáticamente la combinación de grupos de 3-4 personas."
        )

    df = calcular_puntajes_crudos(df)
    df, medias, desviaciones = clasificar_perfiles(df)

    tamanos_grupos = calcular_tamanos_grupos(n_estudiantes)
    print(
        f"Formando {len(tamanos_grupos)} grupos para {n_estudiantes} estudiantes "
        f"con tamaños: {tamanos_grupos}"
    )
    grupos = formar_grupos(df, tamanos_grupos)

    generar_excel_salida(df, grupos, ruta_salida)
    print(f"\nExcel de salida generado en: {ruta_salida}")

    imprimir_resumen(df, grupos, tamanos_grupos)


if __name__ == "__main__":
    main()
